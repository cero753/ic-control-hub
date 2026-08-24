"""Extract IC framework controls from the four source .xlsx files into src/data/controls.json.

Reads the primary sheet of each workbook (ignores *Old archive sheets), maps columns by
header name (robust to the misspelled "Frequancy" header and R2R's duplicated
"Manual / Automated" column that is actually Frequency), and emits one JSON object per control.
Assertion cells are preserved verbatim (P / X / checkmark / blank) for faithful display.
"""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

UP = Path(__file__).resolve().parent.parent
UPLOAD = Path(r"C:\Users\karti\.claude\uploads\a5bfd0f7-4494-4fdf-9100-ec56af94f907")

FILES = [
    ("FSCR", "Financial Statement Closing & Reporting",
     UPLOAD / "c3ea8465-No9_Process_Financial_Statement_Closing_and_Reporting_FSCR.xlsx"),
    ("FA", "Fixed Assets",
     UPLOAD / "d8a77e81-No7_IC_framework__Fixed_Assets.xlsx"),
    ("P2P", "Procure to Pay",
     UPLOAD / "daf41587-No8_IC_framework__P2P_Process.xlsx"),
    ("R2R", "Record to Report",
     UPLOAD / "655c63a8-IC_framework__R2R.xlsx"),
]


def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def key(s):
    return re.sub(r"[^a-z0-9]", "", norm(s).lower())


def pick_sheet(wb):
    """First sheet whose name does not end in 'old'."""
    for ws in wb.worksheets:
        if not ws.title.strip().lower().endswith("old"):
            return ws
    return wb.worksheets[0]


def find_header_row(ws):
    """Row index (1-based) whose cells contain an 'S.No.'-like header."""
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, ws.max_column + 1):
            k = key(ws.cell(r, c).value)
            if k in ("sno", "srno", "serialno") or k.startswith("sno"):
                return r
    return 1


def build_colmap(ws, header_row):
    headers = []  # (col_index, normalized_header)
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(header_row, c).value)
        if v:
            headers.append((c, v))

    colmap = {}

    def match(pred, name):
        for c, h in headers:
            if c in colmap.values():
                continue
            if pred(key(h)):
                colmap[name] = c
                return

    match(lambda k: k.startswith("sno"), "sno")
    match(lambda k: k == "subarea", "subArea")
    match(lambda k: "controlobjective" in k, "controlObjective")
    match(lambda k: k.startswith("risk"), "risk")
    match(lambda k: "keycontrol" in k, "keyControl")
    match(lambda k: "controldescription" in k, "controlDescription")
    match(lambda k: "cutoff" in k, "a_cutoff")
    match(lambda k: k == "accuracy", "a_accuracy")
    match(lambda k: k == "completeness", "a_completeness")
    match(lambda k: "existence" in k, "a_existenceOccurrence")
    match(lambda k: "rights" in k, "a_rightsObligations")
    match(lambda k: "valuation" in k, "a_valuationAllocation")
    match(lambda k: "presentation" in k, "a_presentationDisclosure")
    match(lambda k: "controltype" in k, "controlType")

    # manual/automated + frequency = the columns between last assertion and control type
    start = colmap.get("a_presentationDisclosure") or colmap.get("controlDescription", 0)
    end = colmap.get("controlType", ws.max_column + 1)
    between = [c for c in range(start + 1, end) if c not in colmap.values()
              and norm(ws.cell(header_row, c).value)]
    if between:
        colmap["manualAutomated"] = between[0]
    if len(between) > 1:
        colmap["frequency"] = between[1]
    return colmap


ASSERTION_FIELDS = [
    ("a_cutoff", "cutoff"),
    ("a_accuracy", "accuracy"),
    ("a_completeness", "completeness"),
    ("a_existenceOccurrence", "existenceOccurrence"),
    ("a_rightsObligations", "rightsObligations"),
    ("a_valuationAllocation", "valuationAllocation"),
    ("a_presentationDisclosure", "presentationDisclosure"),
]


def cell(ws, r, colmap, name):
    c = colmap.get(name)
    if not c:
        return ""
    return norm(ws.cell(r, c).value)


def to_bool(v):
    return norm(v).lower().startswith("y")


def requires_request(desc, obj):
    text = (desc + " " + obj).lower()
    return any(w in text for w in ("approv", "authoriz", "authoris"))


controls = []
summary = []
for code, framework, path in FILES:
    wb = load_workbook(path, data_only=True)
    ws = pick_sheet(wb)
    hr = find_header_row(ws)
    colmap = build_colmap(ws, hr)
    n = 0
    for r in range(hr + 1, ws.max_row + 1):
        sub = cell(ws, r, colmap, "subArea")
        obj = cell(ws, r, colmap, "controlObjective")
        desc = cell(ws, r, colmap, "controlDescription")
        if not (sub or obj or desc):
            continue  # skip blank rows
        n += 1
        cid = f"{code}-{n:02d}"
        sub_l = sub.lower()
        request_type = "ledger" if ("chart of account" in sub_l or "chart of account" in obj.lower()) else "generic"
        controls.append({
            "id": cid,
            "framework": framework,
            "frameworkCode": code,
            "sno": n,
            "subArea": sub,
            "controlObjective": obj,
            "risk": cell(ws, r, colmap, "risk"),
            "keyControl": to_bool(cell(ws, r, colmap, "keyControl")),
            "controlDescription": desc,
            "assertions": {out: cell(ws, r, colmap, src) for src, out in ASSERTION_FIELDS},
            "manualAutomated": cell(ws, r, colmap, "manualAutomated"),
            "frequency": cell(ws, r, colmap, "frequency"),
            "controlType": cell(ws, r, colmap, "controlType"),
            "requiresRequest": requires_request(desc, obj),
            "requestType": request_type,
        })
    summary.append(f"{code} ({ws.title}): {n} controls, header row {hr}, cols={sorted(colmap.keys())}")

out_dir = UP / "src" / "data"
out_dir.mkdir(parents=True, exist_ok=True)
out_file = out_dir / "controls.json"
out_file.write_text(json.dumps(controls, indent=2, ensure_ascii=False), encoding="utf-8")

print("\n".join(summary))
print(f"\nTOTAL controls: {len(controls)}")
print(f"requiresRequest: {sum(1 for c in controls if c['requiresRequest'])}")
print(f"ledger requestType: {[c['id'] for c in controls if c['requestType'] == 'ledger']}")
print(f"Wrote {out_file}")
